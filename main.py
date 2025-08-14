import PyPDF2
from pdf2image import convert_from_bytes
import base64
from io import BytesIO
import json
from vox import call_vox_api,get_bearer_token
from openpyxl.drawing.image import Image
from openpyxl import Workbook
import pandas as pd
import requests
from requests.auth import HTTPBasicAuth
import os
from dotenv import load_dotenv
from datetime import datetime
import streamlit as st
from typing import Optional, Dict, Any, List

# Load environment variables
load_dotenv()

vertical_space = 100

def _image_to_base64(image):
    """Convert PIL Image to base64 encoded string"""
    try:
        buffered = BytesIO()
        image.save(buffered, format="PNG")
        img_str = base64.b64encode(buffered.getvalue()).decode('utf-8')
        return img_str
    except Exception as e:
        print(f"Error converting image to base64: {e}")
        return ""

def base64_to_image(base64_str):
    img_data = base64.b64decode(base64_str)
    img = BytesIO(img_data)
    return Image(img)

def extract_annotations(file):
    annotations = {}
    annot_id = 1
    content_coord = {}
    reader = PyPDF2.PdfReader(file)
    num_pages = len(reader.pages)
    name = file.name.split(".pdf")[0]
    file.seek(0)
    file_bytes = file.read()
    page_images = convert_from_bytes(file_bytes)
    
    for i,page in enumerate(reader.pages):
        page_width,page_height = page.mediabox.width, page.mediabox.height
        page_image = page_images[i]
        page_image = page_image.resize((int(page_width), int(page_height)))
        if "/Annots" in page:
            for annot in page.get_object()["/Annots"]:                    
                obj = annot.get_object()
                if  "/Contents" in obj:                        
                    content = obj["/Contents"] if "/Contents" in obj else ""
                    author = obj["/T"] if "/T" in obj else "No Author"
                    coord = obj["/Rect"] or []
                    x0, y0, x1, y1 = coord
                    y0 = float(page_height) - float(y0)
                    y1 = float(page_height) - float(y1)
                    cropped_image = page_image.crop((0, y1-vertical_space, page_width, y0+vertical_space))
                    if tuple(coord) not in content_coord:
                        content_coord[tuple(coord)] = annot_id
                    if '/IRT' in obj:
                        irt_coord = obj['/IRT']['/Rect']
                        irt_annot_id = content_coord[tuple(irt_coord)]
                        annotations[irt_annot_id]['content'].append(content)
                        annotations[irt_annot_id]['author'].append(author)
                    else:
                        coord = [str(coord[i]) for i in range(4)]
                        annotations[annot_id] = {
                            "page": i + 1,
                            "content": [content],
                            "author": [author],
                            "coordinates": coord,
                            "image": _image_to_base64(cropped_image),
                        }
                    annot_id += 1

    annotations = get_defect_nature_llm(annotations)  
    df = create_pandas_df(annotations)    
    return annotations,df

def get_prompts():
    system_prompt = """

    You are a very professional quality analyst, who is expert in UI and content testing. The agency has provided UAT defects where it has images and comments. You will be provided a list of agency authors, their comments and the images.

    You need to follow instruction mentioned in <INSTRUCTION> section.

    <INSTRUCTION>
    The agency has provided UAT defects where it has images and comments. You will be provided a list of agency authors, their comments and the images.
    </INSTRUCTION>
    
    Provide in <DEFINITIONS> section the definitions of the terms used in the task.
    <DEFINITIONS>
    1. Content Defect: The defect which is related to any textual/content changes and issues.
    2. UI Defect: The defect which is related to any UI issues like missing elements, alignment issues, spacing issues, etc.
    3. Change: The defect is called change request if the defect is related to any new feature or change in existing feature.
    4. Bug: The defect is called bug if the defect is related to any existing feature which is not working as expected.
    </DEFINITIONS>

    Your tasks are mentioned in <TASKS> section.
    <TASKS>
    1. Carefully read the images and comments and categorize their nature as either "Content Defect" or "UI Defect".
    2. Carefully read the images and comments and categorize their type as either "Change" or "Bug".
    </TASKS>

    Provide the output in the format mentioned in <OUTPUT> section.
    <OUTPUT>
    {
        "nature":"Content/UI",
        "type":"Change/Bug"
    }
    </OUTPUT>

    Note: Though there are many things to be respond, but only respond the output in the format mentioned in <OUTPUT> section. Do not respond anything else.
    Do not respond the <INSTRUCTION>, <DEFINITIONS>, <TASKS> and <OUTPUT> sections.
    Do not respond the reasoning and explanation of the output.

    """
    return system_prompt

def get_defect_nature_llm(annotations):
    token = get_bearer_token()
    for k,v in annotations.items():
        content = v['content']
        author = v['author']
        image = v['image']
        system_prompt = get_prompts()
        user_input = f"{author},{content}"
        response = call_vox_api(token, system_prompt, user_input, model_name='anthropic.claude-3-5-sonnet-v2:0',max_tokens=50, temperature=0.7,image=image)

        status = response.get("status")
        if status == 'success':
            result = response.get("result")
            result = json.loads(result)
            annotations[k]['nature'] = result.get("nature") 
            annotations[k]['type'] = result.get("type")
            print('annotion_id:',k)
            print("Completion tokens:", response.get("completion_tokens"))
            print("Total tokens:", response.get("total_tokens"))
        else:
            continue
    return annotations

def create_pandas_df(annotations):
    df = pd.DataFrame(columns=['Annotation ID','Image','Page','Content','Author','Coordinates','Nature','Type'])
    for k,v in annotations.items():
        tempContent = ''
        for content in v['content']:
            tempContent += content + '\n\n'
        tempAuthors = ''
        for author in v['author']:
            tempAuthors += author + '\n\n'
        coord = '\n'.join(v['coordinates'])
        df.loc[-1] = [k, 'data:image/png;base64,'+v['image'], v['page'], tempContent, tempAuthors, coord,  v.get('nature'), v.get('type')]  # adding a row
        df.index = df.index + 1  # shifting index
        df = df.sort_index()  # sorting by index
    df.sort_values(by=['Annotation ID'], inplace=True)
    return df

def export_to_csv(df):
    """ Export the DataFrame to a CSV file """
    print(type(df.to_csv(index=False)))
    return df.to_csv(index=False)

def export_to_excel(df):
    """ Export the DataFrame to an Excel file """
    excel_stream = BytesIO()
    wb = Workbook()
    ws = wb.active
    for i,cols in enumerate(df.columns):
        ws[f"{chr(65+i)}1"] = cols
    # print('*'*20)
    for idx, row in df.iterrows():
        for col_idx, col_name in enumerate(df.columns):
            # print(col_name,'-'*20)
            if col_name == 'Image':
                # print(col_idx,row[col_name].startswith("data:image/png;base64,"))
                if row[col_name].startswith("data:image/png;base64,"):
                    row[col_name] = row[col_name][len("data:image/png;base64,"):]
                img = base64_to_image(row[col_name])
                aspect_ratio = img.width / img.height
                img.width = 100
                img.height = int(100 / aspect_ratio)
                cell = f'{chr(65+col_idx)}{idx + 2}'
                ws.add_image(img, cell)
                continue
            ws.cell(row=idx + 2, column=col_idx+1, value=row[col_name])
        
    wb.save(excel_stream)
    excel_stream.seek(0)    
    return excel_stream

def export_to_json(annotations):
    """ Export the annotations to a JSON file """
    json_data = json.dumps(annotations, indent=4)
    return json_data


# Use Attlassian JIRA API to create function for creating a JIRA issue with appropriate paramaterer

def create_jira_issue(project_key, summary, description, issue_type, parent=None, labels=None):
    """Create a JIRA issue with the given parameters."""
    jira_url = os.getenv('JIRA_URL')
    jira_username = os.getenv('JIRA_USERNAME')
    jira_token = os.getenv('JIRA_API_TOKEN')

    if not jira_url or not jira_username or not jira_token:
        raise ValueError("JIRA credentials are not set in environment variables.")

    auth = HTTPBasicAuth(jira_username, jira_token)
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json"
    }
    if parent:
        data = {
            "fields": {
                "project": {
                    "key": project_key
                },
                "summary": summary,
                "description": description,
                "issuetype": {
                    "name": issue_type
                },
                "parent": {"key": parent}
            }
        }
    else:
        data = {
            "fields": {
                "project": {
                    "key": project_key
                },
                "summary": summary,
                "description": description,
                "issuetype": {
                    "name": issue_type
                }
            }
        }

    if labels:
        data["fields"]["labels"] = labels

    response = requests.post(f"{jira_url}/rest/api/2/issue", json=data, headers=headers, auth=auth)

    if response.status_code == 201:
        return response.json()
    else:
        raise Exception(f"Failed to create JIRA issue: {response.status_code} - {response.text}")


# Use Attlassian JIRA API to create function for attaching images to JIRA issue. The images are in base64 format.
def attach_image_to_jira_issue(issue_key, image_base64, filename):
    print('Attaching image.....')
    print(issue_key,filename)
    """Attach an image to a JIRA issue."""
    jira_url = os.getenv('JIRA_URL')
    jira_username = os.getenv('JIRA_USERNAME')
    jira_token = os.getenv('JIRA_API_TOKEN')
    if not jira_url or not jira_username or not jira_token:
        raise ValueError("JIRA credentials are not set in environment variables.")
    auth = HTTPBasicAuth(jira_username, jira_token)
    headers = {
        "Accept": "application/json",
        "X-Atlassian-Token": "no-check"
    }
    data = {
        "file": (filename, BytesIO(base64.b64decode(image_base64)), "image/png")
    }
    response = requests.post(f"{jira_url}/rest/api/2/issue/{issue_key}/attachments", files=data, headers=headers, auth=auth)
    if response.status_code != 200:
        raise ValueError(f"Failed to attach image to JIRA issue: {response.status_code} - {response.text}")

def export_to_jira(annotations):
    """Export annotations to JIRA issues."""
    project_key = os.getenv('JIRA_PROJECT_KEY')
    current_date = datetime.now().strftime("%Y-%m-%d")
    summary = f"UAT Feedback Analysis - {current_date}"
    total_defects = len(annotations.keys())
    description = f"Analysis of UAT feedback containing {total_defects} defects across UI and Content categories"
    issue = create_jira_issue(project_key, summary, description, "Story")
    print(issue.get('key'),'created successfully')
    ui_defects,content_defects  = [],[]
    for k,v in annotations.items():
        v['defect_id'] = k
        if v.get('nature') == 'UI':
            ui_defects.append(v)
        elif v.get('nature') == 'Content':
            content_defects.append(v)
        else:
            continue

    ui_defect_summary = f'UI Defects - {len(ui_defects)} issues identified'
    content_defect_summary = f'Content Defects - {len(content_defects)} issues identified'
   
    ui_defect_description = f"UI Defects Summary:\n\n"
    for defect in ui_defects:
        ui_defect_description += f"\nDefect ID: {defect['defect_id']}\n"
        ui_defect_description += f"Page: {defect['page']}\n"
        ui_defect_description += f"Content: {', '.join(defect['content'])}\n"
        ui_defect_description += f"Reporter: {', '.join(defect['author'])}\n"
        ui_defect_description += f"Type: {defect.get('type')}\n\n--"

    if len(ui_defects) > 0:
        ui_task_issue = create_jira_issue(project_key, ui_defect_summary, ui_defect_description, "Sub-task", parent=issue.get('key'))
        print(ui_task_issue.get('key'),'created successfully')
        for defect in ui_defects:
            if defect.get('image'):
                attach_image_to_jira_issue(ui_task_issue['key'], defect['image'], f"defect_{defect['defect_id']}_page_{defect['page']}.png")
                print(f"Image for UI defect {defect['defect_id']} attached successfully.")
    
    content_defect_description = f"Content Defects Summary:\n\n"
    for defect in content_defects:
        content_defect_description += f"Defect ID: {defect['defect_id']}\n"
        content_defect_description += f"Page: {defect['page']}\n"
        content_defect_description += f"Content: {', '.join(defect['content'])}\n"
        content_defect_description += f"Reporter: {', '.join(defect['author'])}\n"
        content_defect_description += f"Type: {defect.get('type')}\n\n--"

    if len(content_defects) > 0:
        content_task_issue = create_jira_issue(project_key, content_defect_summary, content_defect_description, "Sub-task", parent=issue.get('key'))
        print(content_task_issue.get('key'),'created successfully')    
        for defect in content_defects:
            if defect.get('image'):
                attach_image_to_jira_issue(content_task_issue['key'], defect['image'], f"defect_{defect['defect_id']}_page_{defect['page']}.png")
                print(f"Image for Content defect {defect['defect_id']} attached successfully.")

    return issue.get('key')